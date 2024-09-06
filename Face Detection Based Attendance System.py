import cv2
from face_encoding import Face  # Ensure face_encoding.py is in the same directory
import openpyxl
import os
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from datetime import datetime, timedelta

# Initialize the face recognizer
custom_encoder = Face()

# Load face encodings from the dataset folder (ensure this path exists and has images)
custom_encoder.load_encoding_images("Dataset/")

# Create or load an Excel file for attendance
excel_file = "attendance.xlsx"

if not os.path.exists(excel_file):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Attendance"
    sheet['A1'] = "Name"
    sheet['B1'] = "Status"
    sheet['C1'] = "Time Started"
else:
    workbook = openpyxl.load_workbook(excel_file)
    sheet = workbook.active

# Set for unique names to avoid duplicates in attendance
unique_names = set()
# Dictionary to track the time when attendance started for each student
attendance_start_time = {}
# Time to check if one minute has passed
attendance_duration = timedelta(minutes=1)

# Start capturing video from webcam
cap = cv2.VideoCapture(0)

while True:
    ret, frame = cap.read()
    
    if not ret:
        print("Failed to capture frame from camera. Exiting...")
        break

    # Detect faces and get the recognized names
    face_areas, recognized_names = custom_encoder.detect_known_faces(frame)

    current_time = datetime.now()

    # Display the recognized faces and update attendance
    for face_coords, found_name in zip(face_areas, recognized_names):
        t, r, b, l = face_coords  # Extract face coordinates
        
        if found_name not in unique_names:
            unique_names.add(found_name)
            attendance_start_time[found_name] = current_time
            
            # Find the last row of the sheet and add new entry
            new_row = sheet.max_row + 1
            sheet[f'A{new_row}'] = found_name
            sheet[f'B{new_row}'] = "Present"
            sheet[f'C{new_row}'] = current_time.strftime("%Y-%m-%d %H:%M:%S")
            
            # Save the workbook after each new entry
            workbook.save(excel_file)

        # Check if one minute has passed for the current student
        if found_name in attendance_start_time:
            if current_time - attendance_start_time[found_name] >= attendance_duration:
                # Update attendance status to "Completed" or similar
                for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=1):
                    if row[0].value == found_name:
                        sheet[f'B{row[0].row}'] = "Completed"
                        workbook.save(excel_file)
                        del attendance_start_time[found_name]  # Remove the student from the tracking list
                        break

        # Draw a rectangle around the face and display the name
        cv2.putText(frame, found_name, (l, t - 10), cv2.FONT_HERSHEY_DUPLEX, 1, (0, 0, 200), 2)
        cv2.rectangle(frame, (l, t), (r, b), (0, 0, 200), 4)

    # Display the result frame
    cv2.imshow("Face Recognition Attendance System", frame)

    # Break the loop if 'q' is pressed
    if cv2.waitKey(3) & 0xFF == ord('q'):
        break

# Release the webcam and close windows
cap.release()
cv2.destroyAllWindows()

# Save the final attendance log to the Excel file
workbook.save(excel_file)
workbook.close()

def send_email(subject, body, to_email, attachment_path):
    # Email credentials
    from_email = 'ankushkumargenai@gmail.com'
    password = '900982222'

    try:
        # Setup the MIME
        msg = MIMEMultipart()
        msg['From'] = from_email
        msg['To'] = to_email
        msg['Subject'] = subject

        # Attach the body with the msg instance
        msg.attach(MIMEText(body, 'plain'))

        # Open the file to be sent
        with open(attachment_path, 'rb') as attachment:
            # Instance of MIMEBase and named as part
            part = MIMEBase('application', 'octet-stream')
            part.set_payload(attachment.read())

            # Encode as base64
            encoders.encode_base64(part)

            # Add header
            part.add_header('Content-Disposition', f'attachment; filename= {attachment_path}')

            # Attach the instance 'part' to instance 'msg'
            msg.attach(part)

            # Create SMTP session
            with smtplib.SMTP('smtp.gmail.com', 587) as server:
                server.starttls()  # Enable security
                server.login(from_email, password)  # Login with email and password
                text = msg.as_string()
                server.sendmail(from_email, to_email, text)
                print("Email sent successfully!")

    except Exception as e:
        print(f"Failed to send email. Error: {e}")

# Send the Excel file via email
send_email(
    subject="Attendance Report",
    body="Please find the attached attendance report.",
    to_email="en21cs301109@medicaps.ac.in",
    attachment_path=excel_file
)
